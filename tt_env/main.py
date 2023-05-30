# Standard library imports
import os
import logging
from csv import reader

# Third-party library imports
from PyQt6.QtWidgets import QMainWindow, QFileDialog, QApplication, QDialog
from PyQt6.QtGui import QAction
from openpyxl import load_workbook
from requests import request

# Local application imports
import config as cfg
from PyQt6 import uic


# log setup
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
formatter = logging.Formatter('%(levelname)s:%(name)s:%(asctime)s:%(message)s')
file_handler = logging.FileHandler('repository.log')
file_handler.setFormatter(formatter)
logger.addHandler(file_handler)


class MyGui(QMainWindow):

    def __init__(self):
        super(MyGui, self).__init__()
        uic.loadUi('tt.ui', self)
        self.show()
        self.setWindowTitle('Ticker Track')
        self.active_list.addItems(cfg.ACTIVE_TICKERS[:])
        self.inactive_list.addItems(cfg.INACTIVE_TICKERS[:])
        self.path_line_edit.setText(cfg.BOOK_NAME)
        self.single_radio.setChecked(True)
        self.log_msg = ''
        self.file_path = self.path_line_edit.text()
        self.dates_to_prices = {} # date:[(ticker,price)]
        self.path_line_edit.textChanged.connect(self.ui_refresh)
        self.add_button.clicked.connect(lambda: self.add((self.ticker_line_edit.text().upper())))
        self.toggle_button.clicked.connect(lambda: self.toggle(self.ticker_line_edit.text().upper()))
        self.delete_button.clicked.connect(lambda: self.delete(self.ticker_line_edit.text().upper()))
        self.reset_date_button.clicked.connect(self.load_date)
        self.browse_button.clicked.connect(self.browse)
        self.run_button.clicked.connect(self.run)
        self.single_radio.toggled.connect(self.toggle_date_range)
        self.range_radio.toggled.connect(self.toggle_date_range)
        self.ui_refresh()
        self.load_date()
        self.toggle_date_range()

    '''
        # create a menu bar and add an "About" action to it
        menu_bar = self.menuBar()
        about_action = QAction("About", self)
        about_action.triggered.connect(self.show_about_dialog)
        menu_bar.addAction(about_action)
    
    def show_about_dialog(self):
        about_dialog = QDialog(self)
        about_dialog.setWindowTitle("About My Program")
        about_dialog.setModal(True)
        # set the size and position of the dialog here if needed
        about_dialog.exec()
    '''

    # enable or disable run button based on file path
    def toggle_run_button(self):
        if self.file_path:
            self.run_button.setEnabled(True)
        else:
            self.run_button.setEnabled(False)

    # toggle date range ui widget visibility
    '''def toggle_date_range(self):
        if self.single_radio.isChecked():
            self.year_range_end.setVisible(False)
            self.month_range_end.setVisible(False)
            self.day_range_end.setVisible(False)
            self.date_div_3.setVisible(False)
            self.date_div_4.setVisible(False)
            self.from_label.setVisible(False)
        elif self.range_radio.isChecked():
            self.year_range_end.setVisible(True)
            self.month_range_end.setVisible(True)
            self.day_range_end.setVisible(True)
            self.date_div_3.setVisible(True)
            self.date_div_4.setVisible(True)
            self.from_label.setVisible(True)'''
    
    def toggle_date_range(self):
        date_string = f'{self.year_line_edit.text()}-{self.month_line_edit.text()}-{self.day_line_edit.text()}'
        from_string = f'{self.year_range_end.text()}-{self.month_range_end.text()}-{self.day_range_end.text()}'

        date = cfg.datetime.strptime(date_string, '%Y-%m-%d')

        if self.single_radio.isChecked():
            self.year_range_end.setVisible(False)
            self.month_range_end.setVisible(False)
            self.day_range_end.setVisible(False)
            self.date_div_3.setVisible(False)
            self.date_div_4.setVisible(False)
            self.from_label.setVisible(False)
        elif self.range_radio.isChecked():
            if from_string == '--':
                # populate "from" date fields with values from "date" fields
                self.year_range_end.setText(self.year_line_edit.text())
                self.month_range_end.setText(self.month_line_edit.text())
                self.day_range_end.setText(self.day_line_edit.text())

            from_string = f'{self.year_range_end.text()}-{self.month_range_end.text()}-{self.day_range_end.text()}'
            from_date = cfg.datetime.strptime(from_string, '%Y-%m-%d')

            if from_date is not None and from_date > date:
                # clear the "From" field
                self.year_range_end.clear()
                self.month_range_end.clear()
                self.day_range_end.clear()
                # show an error message or take any other appropriate action
                self.log_msg = 'Error: "From" date cannot be after the "Date" date.'
                logger.error('Error: "From" date cannot be after the "Date" date.')
            else:
                self.year_range_end.setVisible(True)
                self.month_range_end.setVisible(True)
                self.day_range_end.setVisible(True)
                self.date_div_3.setVisible(True)
                self.date_div_4.setVisible(True)
                self.from_label.setVisible(True)

    # init last closing day on startup or reset
    def load_date(self):
        self.year_line_edit.setText(cfg.YEAR)
        self.month_line_edit.setText(cfg.MONTH)
        self.day_line_edit.setText(cfg.DAY)
        self.year_range_end.setText(cfg.YEAR)
        self.month_range_end.setText(cfg.MONTH)
        self.day_range_end.setText(cfg.DAY)

    # get file path from user selection and show user selection in UI
    def browse(self):
        file_name = QFileDialog.getOpenFileName(self, "Open File", "")
        self.path_line_edit.setText(file_name[0]) 

    # store current file path in config
    def save_file_path(self):
        cfg.config.set('FILE', 'book', self.file_path)
        with open('tt_config.ini', 'w') as configfile:
            cfg.config.write(configfile)
    
    # update UI
    def ui_refresh(self):
        # update and sort lists
        self.active_list.clear()
        self.inactive_list.clear()
        cfg.ACTIVE_TICKERS.sort()
        cfg.INACTIVE_TICKERS.sort()
        self.active_list.addItems(cfg.ACTIVE_TICKERS)
        self.inactive_list.addItems(cfg.INACTIVE_TICKERS)
        # update text elements
        self.console_message.setText(self.log_msg)
        self.console_message.setToolTip(self.log_msg)
        self.path_line_edit.setToolTip(self.path_line_edit.text())
        self.file_path = self.path_line_edit.text()
        # update functions
        self.save_file_path()
        self.toggle_run_button()

    # validate and add ticker input to active list
    def add(self, input):
        my_dict = {46: None, 63: None}
        ticker = input.translate(my_dict)
        if ticker and not ticker.isspace():
            url = f'https://api.marketdata.app/v1/stocks/quotes/{ticker}/?token={cfg.KEY}'
            response = request("GET", url)
            # check API for ticker data
            try:
                if bool(response.json()['s'] == 'ok'):
                    if ticker in cfg.ACTIVE_TICKERS:
                        self.log_msg = f'"{ticker}" already in active tickers.'
                        logger.info(f'"{ticker}" already in active tickers.')
                        self.ui_refresh()
                    elif ticker not in cfg.ACTIVE_TICKERS and ticker not in cfg.INACTIVE_TICKERS:
                        cfg.ACTIVE_TICKERS.append(ticker)
                        cfg.update_cfg()
                        self.log_msg = f'"{ticker}" added to active tickers.'
                        logger.info(f'"{ticker}" added to active tickers.')
                        self.ui_refresh()
                    else:
                        self.log_msg = f'"{ticker}" already in inactive tickers.'
                        logger.info(f'"{ticker}" already in inactive tickers.')
                        self.ui_refresh()
                else:
                    self.log_msg = f'API Response: {response.json()["errmsg"]}, Ticker: "{ticker}"'
                    logger.error(f'API Response: {response.json()["errmsg"]}, Ticker: "{ticker}"')
                    self.ui_refresh()
            except:
                self.log_msg = f'Exception: {response.json()["errmsg"]}'
                logger.error(f'Exception: {response.json()["errmsg"]}')
                self.ui_refresh()
        else:
            self.log_msg = 'No Ticker provided.'
            logger.info('No Ticker provided.')
            self.ui_refresh()

    # move ticker between active and inactive lists
    def toggle(self, ticker):
        if ticker and not ticker.isspace():
            if ticker in cfg.ACTIVE_TICKERS:
                cfg.INACTIVE_TICKERS.append(ticker)
                cfg.ACTIVE_TICKERS.remove(ticker)
                cfg.update_cfg()
                self.log_msg = f'"{ticker}" moved to inactive tickers.'
                logger.info(f'"{ticker}" moved to inactive tickers.')
                self.ui_refresh()
            elif ticker in cfg.INACTIVE_TICKERS:
                cfg.ACTIVE_TICKERS.append(ticker)
                cfg.INACTIVE_TICKERS.remove(ticker)
                cfg.update_cfg()
                self.log_msg = f'"{ticker}" moved to active tickers.'
                logger.info(f'"{ticker}" moved to active tickers.')
                self.ui_refresh()
            else:
                self.log_msg = f'"{ticker}" ticker has not yet been added.'
                logger.info(f'"{ticker}" ticker has not yet been added.')
                self.ui_refresh()
        else:
            self.log_msg = 'No Ticker provided.'
            logger.info('No Ticker provided.')
            self.ui_refresh()

    # remove ticker from either list
    def delete(self, ticker):
        if ticker and not ticker.isspace():
            if ticker in cfg.ACTIVE_TICKERS or ticker in cfg.INACTIVE_TICKERS:
                try:
                    cfg.ACTIVE_TICKERS.remove(ticker)
                except:
                    cfg.INACTIVE_TICKERS.remove(ticker)
                finally:
                    cfg.update_cfg()
                    self.log_msg = f'"{ticker}" has been successfully removed.'
                    logger.info(f'"{ticker}" has been successfully removed.')
                    self.ui_refresh()
            else:
                self.log_msg = f'Removal unsuccessful. "{ticker}" not found.'
                logger.info(f'Removal unsuccessful. "{ticker}" not found.')
                self.ui_refresh()
        else:
            self.log_msg = 'No Ticker provided.'
            logger.info('No Ticker provided.')
            self.ui_refresh()

    # process GET requests
    def process_data(self):
        date_string = f'{self.year_line_edit.text()}-{self.month_line_edit.text()}-{self.day_line_edit.text()}'
        from_string = f'{self.year_range_end.text()}-{self.month_range_end.text()}-{self.day_range_end.text()}'
        date = cfg.datetime.strptime(date_string, '%Y-%m-%d')
        from_date = cfg.datetime.strptime(from_string, '%Y-%m-%d')
        if self.single_radio.isChecked():
            dates = [date_string]
        elif self.range_radio.isChecked():
            if from_date > date:
                # show an error message or take any other appropriate action
                self.log_msg = 'Error: "From" date cannot be after the "Date" date.'
                logger.error('Error: "From" date cannot be after the "Date" date.')
                self.ui_refresh()
                return
            else:
                delta = date - from_date
                dates = [cfg.datetime.strftime(from_date + cfg.timedelta(days=i), '%Y-%m-%d') for i in range(delta.days + 1)]
        # create list of dates from UI
        if self.single_radio.isChecked():
            dates = cfg.generate_date_range(date_string, date_string)
        elif self.range_radio.isChecked():
            dates = cfg.generate_date_range(from_string, date_string)
        # create variable to store the last numerical price
        last_price = None
        # GET requests to API 
        for t in cfg.ACTIVE_TICKERS: 
            # initialize last price for current ticker
            last_price = None
            for date in dates:
                url = f'https://api.marketdata.app/v1/stocks/candles/D/{t}?limit=1&date={date}&headers=false&format=csv&columns=c&token={cfg.KEY}' 
                response = request("GET", url)
                # check if response is a valid numerical price
                if cfg.isfloat(response.text.strip()):
                    price = float(response.text.strip())
                    # update last numerical price for current ticker
                    last_price = price
                    self.dates_to_prices.setdefault(date, {})[t] = price
                else:
                    # if last numerical price is not available, go back one day at a time up to 7 days to find a valid price
                    if last_price is None:
                        found_valid_price = False
                        days_back = 0
                        while not found_valid_price and days_back < 7:
                            yesterday = cfg.datetime.strptime(date, '%Y-%m-%d') - cfg.timedelta(days=1)
                            date = cfg.datetime.strftime(yesterday, '%Y-%m-%d')
                            url = f'https://api.marketdata.app/v1/stocks/candles/D/{t}?limit=1&date={date}&headers=false&format=csv&columns=c&token={cfg.KEY}' 
                            response = request("GET", url)
                            if cfg.isfloat(response.text.strip()):
                                price = float(response.text.strip())
                                last_price = price
                                self.dates_to_prices.setdefault(date, {})[t] = price
                                found_valid_price = True
                                print(f'Found valid price {price} for {t} on {date}')
                            else:
                                logger.error(f'API Response: {response.text.strip()} for "{t}" on "{date}"')
                                days_back += 1
                                print(f'Could not find valid price for {t} on {date}, trying {days_back} days back')
                        # check if a valid price was found and add to dictionary
                        if last_price is not None:
                            self.dates_to_prices.setdefault(date, {})[t] = last_price
                            print(f'Using last valid price {last_price} for {t} on {date}')
                        else:
                            print(f'No valid price found for {t} on {date}')
                    else:
                        self.dates_to_prices.setdefault(date, {})[t] = last_price
                        print(f'Using last valid price {last_price} for {t} on {date}')
        print(self.dates_to_prices)

    # file validation
    def write_test(self):
        book = load_workbook(self.file_path)
        book.save(self.file_path)

    # write data to Excel file
    def write_data(self, dict):
        # load the workbook and delete Sheet2 if it exists
        book = load_workbook(self.file_path)
        if 'Sheet2' in book.sheetnames:
            del book['Sheet2']
            book.save(self.file_path)
        # create Sheet2 and copy the data from Sheet1
        sheet1 = book['Sheet1']
        if 'Sheet2' not in book.sheetnames:
            sheet2 = book.create_sheet('Sheet2')
            book.save(self.file_path)
            for row in sheet1.rows:
                values = [cell.value for cell in row]
                sheet2.append(values)
                book.save(self.file_path)
        # clear all rows in Sheet1
        sheet1.delete_rows(1, sheet1.max_row)
        book.save(self.file_path)
        # write newest data to Sheet1               
        book = load_workbook(self.file_path)
        sheet = book['Sheet1']
        row = 1
        for date_idx, date in enumerate(dict):
            idx_row = 1
            for ticker in cfg.ACTIVE_TICKERS:
                price = dict.get(date, {}).get(ticker)
                if price is not None:
                    sheet.cell(row=row, column=1).value = date
                    sheet.cell(row=row, column=2).value = ticker
                    sheet.cell(row=row, column=4).value = price
                    if date_idx > 0:
                        offset = date_idx * 6
                        sheet.cell(row=idx_row, column=1 + offset).value = date
                        sheet.cell(row=idx_row, column=2 + offset).value = ticker
                        sheet.cell(row=idx_row, column=4 + offset).value = price
                        idx_row += 1
                    row += 1
        book.save(self.file_path)
        # get the state of the radio buttons in the UI
        is_single_date = self.single_radio.isChecked()
        is_date_range = self.range_radio.isChecked()
        if is_single_date:
            self.log_msg = self.add_timestamp_to_log(f'Saved results for single date to {self.file_path}')
            logger.info(f'Saved results for single date to {self.file_path}')
        elif is_date_range:
            self.log_msg = self.add_timestamp_to_log(f'Saved results for date range to {self.file_path}')
            logger.info(f'Saved results for date range to {self.file_path}')

    def run(self):
        # get the file path
        self.file_path = self.path_line_edit.text()
        # test if file open or exists
        try:
            self.write_test()
            # pull API data
            try:
                self.process_data()
                self.write_data(self.dates_to_prices)
            except Exception as e:
                self.log_msg = f'An error occurred: {e}'
                logger.error(f'An error occurred: {e}', exc_info=True)
                self.ui_refresh()
                return
            finally:
                # clear dict for next run if program kept open
                self.dates_to_prices = {}
                self.ui_refresh()
        except FileNotFoundError:
            self.log_msg = f'File not found: {self.file_path}'
            logger.error(f'File not found: {self.file_path}')
            self.ui_refresh()
            return
        except PermissionError:
            self.log_msg = f'Permission denied: Unable to write to or access the file: {self.file_path}'
            logger.error(f'Unable to write to or access the file: {self.file_path}')
            self.ui_refresh()
            return

    def add_timestamp_to_log(self, log_msg):
        timestamp = cfg.datetime.now().strftime('%H:%M:%S')
        return f'[{timestamp}] {log_msg}'

def main():
    app = QApplication([])
    window = MyGui()
    app.exec()

if __name__ == '__main__':
    main()